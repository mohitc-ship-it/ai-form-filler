import pandas as pd
import json
from io import StringIO
from typing import List, Optional
from pydantic import BaseModel, Field, create_model

from llm import llm_query, llm_structured
from noPklRetrieval import rag, create_retriever
from ragAnything import store_document, query_rag

# === Pydantic Schemas ===

class FieldValue(BaseModel):
    name: str
    value: Optional[str] = None
    confidence: Optional[float] = Field(default=None, description="Confidence in filled value")
    reason: Optional[str] = Field(default=None, description="Explanation of how value was inferred")


class TableData(BaseModel):
    name: str
    headers: List[str]
    rows: List[dict]


# === STEP 1: Read Excel as CSV ===
def read_excel_as_csv(excel_path: str) -> str:
    """Reads Excel and converts it to CSV text (preserving layout, no column numbers)."""
    df = pd.read_excel(excel_path, header=None)
    csv_text = df.to_csv(index=False, header=False)
    return csv_text



# === STEP 2: Enrich extract JSON using RAG + LLM Structured ===
async def enrich_extract_with_rag(extract_json: dict) -> dict:
    """
    Use RAG and structured LLM outputs to fill missing values
    in both fields and tables.
    """
    retriever = create_retriever("./chroma_db_compliance2", "surveys_lease_data")
    filled_data = {"fields": [], "tables": []}

    # --- Fill missing field values ---
    for field in extract_json["data"].get("fields", []):
    # for field in []:
        name = field.get("name")
        print("field is, ", field)
        if not name:
            continue

        if not field.get("value"):  # Only enrich empty fields
            query = f"""Based on the context What is the best possible value for the field '{name}' based on document context?"""
            print("query before rag is ", query)
            query = field.get("query", query)
            print("query after condition is ", query)
            # rag_context = rag(query, retriever.vectorstore)
            rag_context = await query_rag(query)


            structured_field = llm_structured(
                query=f"{query}'. Context:\n{rag_context}",
                output_schema=FieldValue
            )

            print(f"🧩 Filled field: {structured_field}")
            if structured_field.value is None:
                field["value"] = None
            else:
                field["value"] = structured_field.value.replace(",", " ")


            field["confidence"] = structured_field.confidence or 0.9
            field["reason"] = structured_field.reason or "Filled by RAG+LLM"

     
        filled_data["fields"].append(field)

    # --- Fill missing tables ---
    for table in extract_json["data"].get("tables", []):
        headers = table.get("headers", [])
        table_name = table.get("name", "Unknown Table")

        # if not headers:
        #     print(f"⚠️ Skipping table '{table_name}' — no headers found.")
        #     continue

        if not table.get("rows"):
            query = f"Provide all rows for table '{table_name}' with headers: {headers}. Use prior document context."
            # rag_context = rag(query, retriever.vectorstore)
            rag_context = await query_rag(query)

            # Create dynamic Pydantic model for rows (each header → str)
            field_definitions = {h: (Optional[str], None) for h in headers}
            DynamicRow = create_model("DynamicRow", **field_definitions)

            class DynamicTable(BaseModel):
                name: str
                rows: List[DynamicRow]

            query = f"Context:\n{rag_context}\nNow fill rows for table '{table_name}' with headers {headers}."
            query = field.get("query", query)

            structured_table = llm_structured(
                query=query,
                output_schema=DynamicTable
            )
            print("field ", structured_table)

            rows_output = [r.model_dump() for r in structured_table.rows]
            print(f"📊 Filled table '{table_name}' with {len(rows_output)} rows.")
            table["rows"] = rows_output

        filled_data["tables"].append(table)

    extract_json["data"].update(filled_data)
    return extract_json

class CSVModel(BaseModel):
    csv_text: str

# === STEP 3: Use LLM to merge enriched JSON into CSV ===
def fill_csv_with_enriched_json(csv_text: str, enriched_json: dict) -> str:
    """LLM merges structured enriched JSON into the original CSV template."""
    json_text = json.dumps(enriched_json, indent=2)

    prompt = f"""
    You are a precise data assistant that edits CSV files.

    You are given:
    1️⃣ A **CSV template** showing the exact layout, spacing, headers, and empty cells.
    2️⃣ A **JSON object** containing extracted and enriched field/table values.

    Your job:
    - Fill only the missing or empty cells in the CSV using the matching data from the JSON.
    - Do NOT delete, add, rename, or reorder any rows or columns.
    - Keep every header, comma, line break, and blank field in exactly the same place.
    - Preserve the number of columns on every line.
    - When inserting a value, place it in the correct existing column, respecting CSV alignment.
    - If data is not found in JSON, leave the cell blank.
    - If JSON provides additional data not represented in the CSV template, ignore it.
    - Make sure the final CSV has **the exact same number of lines and commas** as the original.
    - Do not include explanations, comments, or any extra text before or after the CSV.
    -don't add any index or extra lines - keep exact same lines except when adding rows for tables from data.
    remove commans from data which we are getting as we will be converting output csv to excel
    -overall make data such that it can be directly converted to excel without any issues like multiline cells or misaligned columns

    Output:
    👉 Only the completed CSV text — properly comma-delimited, line by line.
    Soltution to this code will be used in csv to excel converstion, so make sure output's comma (,) or asterists and other puncutations are handled properly or comma removed.


    CSV TEMPLATE:
    {csv_text}

    ENRICHED JSON:
    {json_text}
    """


    result = llm_structured(prompt,CSVModel)

    # Clean extra text if any
    # if "CSV View" in result:
    #     result = result.split("CSV View of last chunk:")[-1].strip()

    print('final csv ', result.csv_text)

    return result.csv_text


# === STEP 4: Convert final CSV to Excel ===
# def save_filled_csv_to_excel(csv_text: str, output_excel_path: str):
#     """Convert LLM-filled CSV text into a proper Excel file."""
#     df = pd.read_csv(StringIO(csv_text))
#     df.to_excel(output_excel_path, index=False)
#     print(f"✅ Filled Excel saved to {output_excel_path}")
import pandas as pd
from io import StringIO

import pandas as pd
import csv
from io import StringIO

import pandas as pd
import csv
from io import StringIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import StringIO
import csv
from copy import copy


def save_filled_csv_to_excel_using_template(
    filled_csv_text: str,
    input_excel_path: str,
    output_excel_path: str,
    old_csv_text: str = None
):
    """
    Fills Excel template with data from filled CSV while preserving formatting.
    - If a table-like section has no headers, insert its rows just below the section name.
    - Skips writing or clearing into merged cells (read-only in openpyxl).
    """

    # === Step 1: Load Excel template ===
    wb = load_workbook(input_excel_path)
    ws = wb.active
    original_max_row = ws.max_row
    max_col = ws.max_column

    # === Step 2: Read CSV ===
    new_rows = list(csv.reader(StringIO(filled_csv_text)))
    csv_row_count = len(new_rows)

    # === Step 3: Helper — Detect header presence ===
    def looks_like_header(row):
        if not row:
            return False
        non_empty = [v for v in row if v.strip()]
        if not non_empty:
            return False
        alpha_count = sum(v.strip().isalpha() for v in non_empty)
        return alpha_count / len(non_empty) >= 0.6

    # === Step 4: Write CSV rows into Excel ===
    r_excel = 1
    while r_excel <= csv_row_count and r_excel <= original_max_row:
        csv_row = new_rows[r_excel - 1]
        if not csv_row or all(not v.strip() for v in csv_row):
            r_excel += 1
            continue

        # Detect a section name with no headers below
        if (
            len(csv_row) == 1
            or (csv_row[0].strip() and all(not v.strip() for v in csv_row[1:]))
        ):
            section_name = csv_row[0].strip()
            next_row = new_rows[r_excel] if r_excel < csv_row_count else None
            if next_row and not looks_like_header(next_row):
                # collect following rows as bullet items
                items = []
                j = r_excel + 1
                while j < csv_row_count:
                    row = new_rows[j]
                    if not row or all(not v.strip() for v in row):
                        break
                    if len(row) == 1 and row[0].strip().endswith(":"):
                        break
                    items.append(row)
                    j += 1

                ws.cell(row=r_excel, column=1, value=section_name)
                insert_at = r_excel + 1
                for item_row in items:
                    ws.cell(row=insert_at, column=2, value=" ".join(item_row))
                    insert_at += 1
                r_excel = insert_at
                continue

        # Normal row write
        for c in range(min(max_col, len(csv_row))):
            val = csv_row[c]
            cell = ws.cell(row=r_excel, column=c + 1)
            if type(cell).__name__ == "MergedCell":
                continue
            cell.value = val
        r_excel += 1

    # === Step 5: Clear/append for length differences (skip merged cells) ===
    if csv_row_count < original_max_row:
        for r_clear in range(csv_row_count + 1, original_max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r_clear, column=c)
                if type(cell).__name__ == "MergedCell":
                    continue  # <-- FIX: skip read-only merged cells
                cell.value = None
    elif csv_row_count > original_max_row:
        for r_append in range(original_max_row, csv_row_count):
            ws.append(new_rows[r_append])

    # === Step 6: Preserve column widths and styles ===
    template_wb = load_workbook(input_excel_path)
    template_ws = template_wb.active

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        if letter in template_ws.column_dimensions:
            ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width

    for row in template_ws.iter_rows():
        for cell in row:
            if cell.row <= ws.max_row and cell.column <= ws.max_column:
                target = ws.cell(row=cell.row, column=cell.column)
                if cell.has_style:
                    target.font = copy(cell.font)
                    target.border = copy(cell.border)
                    target.fill = copy(cell.fill)
                    target.number_format = copy(cell.number_format)
                    target.protection = copy(cell.protection)
                    target.alignment = copy(cell.alignment)

    # === Step 7: Save ===
    wb.save(output_excel_path)
    print(f"✅ Excel filled safely and saved at {output_excel_path}")
    return output_excel_path

# === STEP 5: Main pipeline ===
async def run_excel_filling_pipeline(input_excel_path: str, extract_json: dict,csv_text) -> str:
    print("📥 Reading Excel...")
    if not csv_text:
        csv_text = read_excel_as_csv(input_excel_path)

    print("🔍 Enriching extracted JSON with RAG + LLM structured output...")
    enriched_json = await enrich_extract_with_rag(extract_json)

    print("🧠 Filling CSV using enriched JSON + LLM...")
    filled_csv = fill_csv_with_enriched_json(csv_text, enriched_json)
#     filled_csv = """,,,,,,,,,,,,2024-09-30
# ,,,,,,,,,,,,
# ,Investment Summary,,,,,,,,,,,
# ,Address,"4801 West 77th Street",,,,,,,,,,
# ,"City, State","Burbank, Illinois",,,,,,,,,,
# ,Tenant,"GMRI, INC. (Olive Garden #1179)",,,,,,,,,,
# ,Annual Rent,"$122,984.40 (effective Feb 1, 2026)",,,,,,,,,,
# ,Lease Expiration,"January 31, 2036 (Base Term)",,,,,,,,,,
# ,Term Remaining,"137 months (as of September 2024)",,,,,,,,,,
# ,Building Square Footage,"10,640 SF",,,,,,,,,,
# ,Parcel Size (AC),"0.52 acres",,,,,,,,,,
# ,Lease Type,"NNN (Triple Net) - Tenant pays all taxes, insurance, utilities, and 5.33% of CAM",,,,,,,,,,
# ,Landlord Responsibilities,"Repave parking lot on or before September 30, 2024; Exterior Common Area maintenance",,,,,,,,,,
# ,Branch Deposits,"N/A",,,,,,,,,,
# ,,,,,,,,,,,,
# ,,,,,,,,,,,,
# ,,,,,,,,,,,,
# ,Rent Schedule,,,,,,,,,,,
# ,Tenant,Square Feet,Commencement Date,Expiration Date,Step Up Date ,Rent,,,,,,
# ,"Olive Garden #1179","10,640","February 1, 2021","January 31, 2026","N/A" ,"$111,804.00",,,,,,
# ,"Olive Garden #1179","10,640","February 1, 2026","January 31, 2031","February 1, 2026" ,"$122,984.40",,,,,,
# ,"Olive Garden #1179","10,640","February 1, 2031","January 31, 2036","February 1, 2031" ,"$135,282.84",,,,,,
# ,"Olive Garden #1179","10,640","February 1, 2036","January 31, 2041","February 1, 2036" ,"$148,811.12",,,,,,
# ,"Olive Garden #1179","10,640","February 1, 2041","January 31, 2046","February 1, 2041" ,"$163,692.23",,,,,,
# ,"Olive Garden #1179","10,640","February 1, 2046","January 31, 2051","February 1, 2046" ,,,,,,,
# ,,,,,,,,,,,,
# ,Strengths,,,,,,,,,,,
# ,"Strong national credit tenant (Darden Restaurants, parent company)",,,,,,,,,,
# ,"Extended term through 2036 with three 5-year extension options through 2051",,,,,,,,,,
# ,"Triple net lease structure minimizes landlord operating expenses",,,,,,,,,,
# ,"Predictable rent escalations of ~10% every 5 years",,,,,,,,,,
# ,"No percentage rent obligation (eliminated in Third Amendment)",,,,,,,,,,
# ,"Protected parking area and Lot 8 restrictions provide operational stability",,,,,,,,,,
# ,"Established location operating since 1989",,,,,,,,,,
# ,,,,,,,,,,,,
# ,Weaknesses,,,,,,,,,,,
# ,"Landlord must complete parking lot repaving by September 30, 2024",,,,,,,,,,
# ,"Extensive tenant CAM audit rights (Landlord pays costs if overcharge >5%)",,,,,,,,,,
# ,"Lot 8 development restrictions limit future flexibility",,,,,,,,,,
# ,"Protected Area covenants restrict property modifications",,,,,,,,,,
# ,"Rent abatement risk if lift station failure causes closure >48 hours",,,,,,,,,,
# ,"Small parcel size (0.52 acres) limits alternative uses",,,,,,,,,,"""

    output_excel = input_excel_path.replace(".xlsx", "_filled.xlsx")
    # save_filled_csv_to_excel(filled_csv, output_excel)
    save_filled_csv_to_excel_using_template(filled_csv, input_excel_path, output_excel)

    print("✅ Pipeline complete.")
    return output_excel, filled_csv


# === Script Entry ===
if __name__ == "__main__":
    input_excel = "userPdfData/refinalfollowupactionrequiredforyourprototypedeve (1)/BOV Template.xlsx"

    with open("extracted_excel_structure.json", "r") as f:
        extract_json = json.load(f)

    output_path,filled_csv = run_excel_filling_pipeline(input_excel, extract_json)
    print("Output file saved at:", output_path)
